# leads/views.py
import json
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404, render, redirect
from .models import Lead
from notifications.models import Notification  # Import Notification from the notifications app
from .forms import LeadForm
from .forms import AssignLeadForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
import csv
from datetime import datetime
from django.utils import timezone
from datetime import timedelta
import openpyxl
from django.db.models import Q
from users.models import Announcement
from django.db.models.functions import TruncDate 
from django.db.models import Count
from django.utils.dateparse import parse_date
from notifications.models import Notification  # Import the Notification model
from django.http import JsonResponse
from .api_integration import fetch_google_ads_data
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.db.models.functions import TruncMonth
from django.db.models import Count
from django.utils import timezone
from .models import Lead
@login_required
def lead_list(request):
    # Get status choices from the Lead model
    statut_choices = Lead._meta.get_field('statut').choices
    
    # Get filters from the request
    statut_filter = request.GET.get('statut', 'all')
    created_at_filter = request.GET.get('crée', None)

    # Count leads for each status
    statut_counts = {
        'nouveau': Lead.objects.filter(statut='nouveau').count(),
        'Contacté': Lead.objects.filter(statut='Contacté').count(),
        'Perdu': Lead.objects.filter(statut='Perdu').count(),
    }
    
    # Query leads and apply filters
    leads = Lead.objects.all()

    # Apply status filter
    if statut_filter != 'all':
        leads = leads.filter(statut=statut_filter)

    # Apply date filter
    if created_at_filter:
        try:
            parsed_date = parse_date(created_at_filter)
            if parsed_date:
                leads = leads.annotate(date=TruncDate('crée')).filter(date=parsed_date)
        except ValueError:
            print('Invalid date format')

    # Query leads that are pending follow-up (not assigned)
    leads_pending_followup = Lead.objects.filter(Assigné__isnull=True)
   
    # Filtrer les leads pris en charge
    leads_pris_en_charge = Lead.objects.filter(Assigné__isnull=False)
    
    # Filtrer les leads non pris en charge
    leads_non_pris_en_charge = Lead.objects.filter(Assigné__isnull=True)
    
    # Filtrer les leads pris en charge et dont le statut est différent de 'Contacté'
    leads_non_contactes = Lead.objects.filter(~Q(statut='Contacté'))
    
    # Compter le nombre de leads pris en charge
    nombre_leads_pris_en_charge = leads_pris_en_charge.count()
    
    # Compter le nombre de leads non pris en charge
    nombre_leads_non_pris_en_charge = leads_non_pris_en_charge.count()
    
    # Compter le nombre de leads pris en charge et non contactés
    nombre_leads_non_contactes = leads_non_contactes.count()

    # Trier les leads par date de création décroissante et limiter à 4
    leads = leads.order_by('-crée')[:4]

    # Indiquer si le bouton "Voir plus" doit être affiché
    show_voir_plus = leads.count() == 4

    # Récupérer les 3 rendez-vous les plus imminents
    now = timezone.now()
    upcoming_appointments = Appointment.objects.filter(date__gte=now).order_by('date')[:3]

    # Calculer le nombre de leads générés chaque mois
    leads_by_month = Lead.objects.annotate(month=TruncMonth('crée')).values('month').annotate(count=Count('id')).order_by('month')

    # Créer une liste de tous les mois de l'année et initialiser les comptes à zéro
    months = []
    counts = []
    notifications = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')
    notifications_count = notifications.count()
    # Crée une liste des mois de l'année en français
    month_names_fr = [
        'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
    ]

    # Crée une liste des mois de l'année
    for i in range(1, 13):
        month = datetime(datetime.now().year, i, 1)
        month_str = month.strftime('%B')  # Juste le mois sans l'année
        month_str_fr = month_names_fr[i - 1]  # Convertir en mois français
        months.append(month_str_fr)
        # Cherche le nombre de leads pour ce mois ou initialise à 0
        count = next((item['count'] for item in leads_by_month if item['month'].strftime('%B') == month_str), 0)
        counts.append(count)
        announcements = Announcement.objects.all().order_by('-created_at')[:4]

    context = {
        'nombre_leads_pris_en_charge': nombre_leads_pris_en_charge,
        'nombre_leads_non_pris_en_charge': nombre_leads_non_pris_en_charge,
        'nombre_leads_non_contactes': nombre_leads_non_contactes,
        'leads': leads,
        'statut_choices': statut_choices,
        'created_at_filter': created_at_filter,
        'statut_counts': statut_counts,
        'leads_pending_followup': leads_pending_followup,
        'show_voir_plus': show_voir_plus,  # Pour savoir si afficher le bouton "Voir plus"
        'upcoming_appointments': upcoming_appointments,  # Ajout des rendez-vous imminents
        'months': months,
        'counts': counts,
        'announcements': announcements, 
        'notifications': notifications,
        'notifications_count': notifications_count, # Ajout des annonces au contexte
    }
    return render(request, 'lead_list.html', context)
   
def faq(request):
    return render(request, 'faq.html')
from django.views.generic import TemplateView

class SupportGuideView(TemplateView):
    template_name = 'support_guide.html'

def lead_detail(request, lead_id):
    lead = Lead.objects.get(id=lead_id)
    return render(request, 'lead_detail.html', {'lead': lead})
from django.shortcuts import render, get_object_or_404, redirect
from .models import Lead, Interaction
from .forms import AppelForm
  # type: ignore # Assurez-vous que cette fonction est correctement définie

@login_required
def interaction_view(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    
    if request.method == "POST":
        form = AppelForm(request.POST)
        if form.is_valid():
            interaction = form.save(commit=False)
            interaction.lead = lead
            interaction.type = 'Appel'  # Explicitly set the type
            interaction.save()
            return redirect('lead_interactions', lead_id=lead.id)
    else:
        form = AppelForm()

    interactions_telephonique = Interaction.objects.filter(lead=lead, type="Appel")
    interactions_email = Interaction.objects.filter(lead=lead, type="Email")

    return render(request, 'lead_interactions.html', {
        'lead': lead,
        'appel_form': form,
        'interactions_telephonique': interactions_telephonique,
        'interactions_email': interactions_email,
    })
import os
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings

def save_call_to_excel(lead, cleaned_data):
    # Définir le chemin du fichier Excel pour tous les leads
    file_path = os.path.join(settings.MEDIA_ROOT, "lead_interactions.xlsx")

    # Vérifier si le fichier Excel existe déjà
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    else:
        # Créer un nouveau fichier Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Ajouter les en-têtes de colonnes
        headers = ["Date", "Durée", "Résumé", "Notes"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

    # Trouver la première ligne vide
    next_row = sheet.max_row + 1

    # Ajouter les données de l'appel
    sheet.cell(row=next_row, column=1, value=cleaned_data['date'])
    sheet.cell(row=next_row, column=2, value=cleaned_data['duree'])
    sheet.cell(row=next_row, column=3, value=cleaned_data['resume'])
    sheet.cell(row=next_row, column=4, value=cleaned_data['notes'])

    # Ajuster la largeur des colonnes
    for col_num in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].auto_size = True

    # Enregistrer le fichier Excel
    workbook.save(file_path)

    return file_path



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Lead
from .forms import LeadForm
from notifications.models import Notification

@login_required
def lead_add(request):
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)
            lead.save()

            # Crée une notification pour chaque utilisateur
            users = User.objects.all()
            for user in users:
                Notification.objects.create(
                    user=user,
                    message=f'Un nouveau lead "{lead.Prénom}" a été ajouté.',
                    type='lead'
                )

            return redirect('lead_actions')
        else:
            print(form.errors)  # Affiche les erreurs du formulaire pour le débogage
    else:
        form = LeadForm()

    return render(request, 'lead_form.html', {'form': form})

def lead_delete(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    if request.method == 'POST':
        lead.delete()
        return redirect('lead_actions')  # Redirection vers la liste des leads après suppression réussie
    return render(request, 'lead_confirm_delete.html', {'lead': lead})
def lead_edit(request, lead_id):
    # Récupérer le lead basé sur l'ID
    lead = get_object_or_404(Lead, id=lead_id)

    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            form.save()
            # Redirection vers la liste des leads après modification réussie
            return redirect('lead_actions')  # Vous pouvez changer 'lead_actions' si nécessaire
    else:
        form = LeadForm(instance=lead)

    return render(request, 'lead_form.html', {'form': form, 'lead': lead})
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Lead
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Lead
import csv
import json
from io import StringIO
def parse_file(file):
    try:
        content = file.read().decode('utf-8')
    except UnicodeDecodeError:
        content = file.read().decode('latin-1')
    
    if file.name.endswith('.csv') or file.name.endswith('.txt'):
        reader = csv.DictReader(content.splitlines())
        headers = reader.fieldnames
        required_fields = {'Prénom', 'Nom', 'Email', 'Télephone', 'source', 'statut', 'Notes', 'Créé', 'Assigné'}
        
        # Vérifiez si les en-têtes du fichier sont présents
        missing_headers = required_fields - set(headers)
        if missing_headers:
            raise ValueError(f"Les en-têtes suivants sont manquants : {missing_headers}")
        
        return list(reader)
    else:
        raise ValueError("Format de fichier non supporté")
def lead_import(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        errors = []

        try:
            data = parse_file(csv_file)
            print("Données importées :", data)  # Ajoutez ceci pour voir les données
        except ValueError as e:
            errors.append(str(e))

        if isinstance(data, list):
            for row in data:
                print("Traitement de la ligne :", row)  # Ajoutez ceci pour voir chaque ligne traitée
                source = row.get('source', '').strip() if row.get('source') else ''
                statut = row.get('statut', '').strip() if row.get('statut') else ''
                assigné_username = row.get('Assigné', '').strip() if row.get('Assigné') else None
                
                # Vérifiez la validité de la source et du statut
                if source not in dict(Lead.SOURCE_CHOICES) or statut not in dict(Lead.STATUS_CHOICES):
                    errors.append(f"Valeur invalide pour source ou statut: {source}, {statut}")
                    continue

                # Si assigné_username est None, ne pas tenter de récupérer l'utilisateur
                assigné_user = None
                if assigné_username:
                    try:
                        assigné_user = User.objects.get(username=assigné_username)
                    except User.DoesNotExist:
                        errors.append(f"L'utilisateur avec le nom d'utilisateur '{assigné_username}' n'existe pas.")
                        continue

                try:
                    Lead.objects.create(
                        Prénom=row.get('Prénom', '').strip() if row.get('Prénom') else '',
                        Nom=row.get('Nom', '').strip() if row.get('Nom') else '',
                        Email=row.get('Email', '').strip() if row.get('Email') else '',
                        Télephone=row.get('Télephone', '').strip() if row.get('Télephone') else '',
                        source=source,
                        statut=statut,
                        notes=row.get('Notes', '').strip() if row.get('Notes') else '',
                        crée=row.get('Créé', '').strip() if row.get('Créé') else None,
                        Assigné=assigné_user
                    )
                except Exception as e:
                    errors.append(f"Erreur lors de l'ajout du lead: {str(e)}")
                    continue

            if not errors:
                
                return redirect('lead_actions')
            else:
                for error in errors:
                    messages.error(request, error)

        else:
            messages.error(request, "Le format du fichier n'est pas reconnu.")
    
    return render(request, 'lead_import.html')




def lead_api_import(request):
    # Code for API integration to import leads automatically
    return HttpResponse("API Integration for lead import")

from django.shortcuts import render

def lead_ac(request):
    # Logique de la vue ici
    return render(request, 'lead_ac.html')

@login_required
def lead_assign(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    if not lead.Assigné and request.user.is_authenticated:
        lead.Assigné = request.user
        lead.save()
    return redirect('lead_list')
@login_required
def lead_actions(request):
    leads = Lead.objects.all()
    
    # Récupérer les paramètres de filtrage
    status_filter = request.GET.get('status', 'all')

    # Affichage des filtres reçus pour le débogage
    print(f"Status Filter: {status_filter}")

    # Appliquer le filtre de statut
    if status_filter and status_filter != 'all':
        leads = leads.filter(statut=status_filter)

    # Calculer le nombre total de leads
    leads_count = leads.count()

    # Calculer les mois et les counts
    leads_by_month = Lead.objects.annotate(month=TruncMonth('crée')).values('month').annotate(count=Count('id')).order_by('month')

    # Crée une liste des mois de l'année en français
    month_names_fr = [
        'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
    ]

    # Créer les listes pour les mois et les comptes
    months = []
    counts = []

    # Crée une liste des mois de l'année
    for i in range(1, 13):
        month = datetime(datetime.now().year, i, 1)
        month_str = month.strftime('%B')  # Mois en anglais
        month_str_fr = month_names_fr[i - 1]  # Convertir en mois français
        months.append(month_str_fr)
        # Cherche le nombre de leads pour ce mois ou initialise à 0
        count = next((item['count'] for item in leads_by_month if item['month'].strftime('%B') == month_str), 0)
        counts.append(count)

    return render(request, 'lead_actions.html', {
        'leads': leads,
        'status_filter': status_filter,
        'leads_count': leads_count,
        'months': months,
        'counts': counts
    })

# Access all leads assigned to a specific user
#assigned_leads = user.assigned_leads.all()
from django.shortcuts import render
# Vue basée sur une classe
from django.views.generic import TemplateView
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.db.models.functions import TruncMonth
from .models import Lead
from datetime import datetime

from django.utils import timezone
from django.views.generic import TemplateView
from leads.models import Lead
from appointments.models import Appointment

class DashboardView(TemplateView):
    template_name = 'lead_list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user
        
        # Récupération des leads en attente de suivi
        context['leads_pending_followup'] = Lead.objects.filter(statut='pending').order_by('créé')[:3]
        
        # Récupération des 3 rendez-vous les plus proches
        now = timezone.now()
        context['upcoming_appointments'] = Appointment.objects.filter(date__gte=now).order_by('date')[:3]
        
        return context



from django.shortcuts import render
from .models import Lead

def dashboard(request):
    # Filtrer les leads non pris en charge et les ordonner par date de création
    leads_pending_followup = Lead.objects.filter(assigné__isnull=True).order_by('crée')[:3]
    context = {
        'leads_pending_followup': leads_pending_followup,
    }
    return render(request, 'dashboard.html', context)

def import_google_ads_leads(request):
    access_token = 'YOUR_GOOGLE_ACCESS_TOKEN'
    data = fetch_google_ads_data(access_token)

    # Exemple d'importation des données
    for item in data.get('results', []):
        Lead.objects.create(
            nom=item.get('ad_group_ad', {}).get('ad', {}).get('name', ''),
            prenom='',
            email='',
            telephone='',
            source='Google Ads',
            statut='Nouveau',
            note='',
            # Ajoute les autres champs nécessaires
        )

    return JsonResponse({'status': 'success', 'data': data})

from django.http import JsonResponse
from .models import Lead

def get_lead_statistics(request):
    tasks_accomplished = Lead.objects.filter(contacted=True).count()
    tasks_not_accomplished = Lead.objects.filter(contacted=False).count()
    
    # Ajoutez des statistiques supplémentaires si nécessaire
    data = {
        'tasks_accomplished': tasks_accomplished,
        'tasks_not_accomplished': tasks_not_accomplished,
        'meetings_upcoming': 5,  # Exemple statique, remplacez par des données réelles
        'emails_pending': 8  # Exemple statique, remplacez par des données réelles
    }
    
    return JsonResponse(data)

from .forms import AppelForm


@login_required
def historique_appels(request, lead_id):
    lead = get_object_or_404(Lead, id=lead_id)
    appels = Interaction.objects.filter(lead=lead, type='appel').order_by('-date')
    return render(request, 'historique_appels.html', {'lead': lead, 'appels': appels})



from django.shortcuts import redirect
from .models import EmailHistory  # Assurez-vous que ce modèle est défini
from django.shortcuts import redirect
from django.utils import timezone
from .models import EmailHistory
# views.py
from django.shortcuts import redirect
from django.utils import timezone
from .models import EmailHistory
# views.py
from django.shortcuts import redirect
from django.utils import timezone
from django.conf import settings
from .models import EmailHistory
from django.core.mail import send_mail

def send_email(request, lead_id):
    if request.method == 'POST':
        email_to = request.POST.get('to')
        email_subject = request.POST.get('subject')
        email_body = request.POST.get('body')

        # Enregistrer l'email dans l'historique
        EmailHistory.objects.create(
            lead_id=lead_id,
            to=email_to,
            subject=email_subject,
            body=email_body,
            sent_at=timezone.now()
        )

        # Envoyer l'email via Gmail (redirection)
        return redirect(f"https://mail.google.com/mail/?view=cm&fs=1&to={email_to}&su={email_subject}&body={email_body}")

    return redirect('error')



    return redirect('error')  # Redirection en cas d'erreur
from django.shortcuts import render
from .models import EmailHistory

# views.py
from django.shortcuts import render, get_object_or_404
from .models import EmailHistory, Lead

def historique_emails(request, lead_id):
    # Récupérer l'objet Lead basé sur lead_id (si nécessaire)
    lead = get_object_or_404(Lead, pk=lead_id)
    # Filtrer les emails associés à ce lead
    emails = EmailHistory.objects.filter(lead_id=lead_id)
    # Passer les emails et l'objet lead au template
    return render(request, 'historique_emails.html', {'emails': emails, 'lead': lead})

def BoiteGmail(request, lead_id):
    return render(request, 'BoiteGmail.html', {'lead_id': lead_id})
